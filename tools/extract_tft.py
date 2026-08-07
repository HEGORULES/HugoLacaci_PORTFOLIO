#!/usr/bin/env python3
"""Extract the TFT workbooks into docs/data/tft.json.

The site renders the sheets as real tables and runs the damage model in
JavaScript, so the spreadsheets are readable and usable without Excel.
Re-run after changing either .xlsx:

    python3 tools/extract_tft.py
"""

import json
import pathlib

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
FILES = ROOT / "docs" / "files"
OUT = ROOT / "docs" / "data" / "tft.json"

# Stat columns, in the order the workbook lists them.
STATS = ["hp", "ad", "ap", "armor", "mr", "as", "crit", "critDmg",
         "omnivamp", "dmgAmp", "durability"]


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def sheet_rows(ws):
    """Trim a sheet to its genuinely populated rectangle."""
    rows = []
    for r in ws.iter_rows():
        vals = [clean(c.value) for c in r]
        while vals and vals[-1] is None:
            vals.pop()
        rows.append(vals)
    while rows and not any(rows[-1]):
        rows.pop()
    lead = min((next((i for i, v in enumerate(r) if v is not None), 99)
                for r in rows if any(v is not None for v in r)), default=0)
    return [r[lead:] for r in rows if any(v is not None for v in r)]


def num(cells, idx):
    """0-based column index -> number, or 0."""
    if idx >= len(cells):
        return 0
    v = clean(cells[idx].value)
    return v if isinstance(v, (int, float)) else 0


def champions(wb):
    """Stats sit in C..M. Column N is empty; the scalings are in O and P,
    which is why the workbook's VLOOKUPs reach for indices 14 and 15."""
    ws = wb["Champions"]
    out = []
    for r in ws.iter_rows(min_row=3):
        name = clean(r[1].value) if len(r) > 1 else None
        if not isinstance(name, str):
            continue
        rec = {"name": name}
        for i, k in enumerate(STATS):
            rec[k] = num(r, 2 + i)          # C..M
        rec["adScaling"] = num(r, 14)        # O
        rec["apScaling"] = num(r, 15)        # P
        if not any(rec[k] for k in STATS):
            continue
        out.append(rec)
    return out


def items(wb_values):
    """Items are formulas over BaseItem, so read the cached values."""
    ws = wb_values["Items"]
    out = []
    for r in ws.iter_rows(min_row=3):
        name = clean(r[1].value) if len(r) > 1 else None
        if not isinstance(name, str):
            continue
        vals = [clean(c.value) for c in r[2:13]]
        if not any(isinstance(v, (int, float)) for v in vals):
            continue
        rec = {"name": name}
        for i, k in enumerate(STATS):
            rec[k] = vals[i] if i < len(vals) and isinstance(vals[i], (int, float)) else 0
        out.append(rec)
    return out


def main():
    calc_f = openpyxl.load_workbook(FILES / "tft-damage-calculator.xlsx", data_only=False)
    calc_v = openpyxl.load_workbook(FILES / "tft-damage-calculator.xlsx", data_only=True)
    traits_v = openpyxl.load_workbook(FILES / "tft-set-16-traits.xlsx", data_only=True)

    data = {
        "champions": champions(calc_f),
        "items": items(calc_v),
        "sheets": {
            "traits": [
                {"name": ws.title, "rows": sheet_rows(ws)}
                for ws in traits_v.worksheets
            ],
            "calculator": [
                {"name": ws.title, "rows": sheet_rows(ws)}
                for ws in calc_v.worksheets if ws.title != "Calculator"
            ],
        },
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    kb = OUT.stat().st_size / 1024
    print(f"  champions: {len(data['champions'])}")
    print(f"  items:     {len(data['items'])}")
    for group, sheets in data["sheets"].items():
        for s in sheets:
            print(f"  {group}/{s['name']}: {len(s['rows'])} rows")
    print(f"wrote {OUT} ({kb:.0f} KB)")


if __name__ == "__main__":
    main()
