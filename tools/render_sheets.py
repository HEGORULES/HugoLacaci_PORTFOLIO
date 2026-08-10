#!/usr/bin/env python3
"""Extract both .xlsx into docs/data/tft-sheets.json as a live model.

Every cell keeps whatever it actually holds, a formula stays a formula,
a literal stays a literal, plus the styling that makes the sheet
readable. The browser renders the grid and evaluates the formulas, so
the workbook can be edited and recalculated rather than just looked at.

    python3 tools/render_sheets.py
"""

import colorsys
import json
import pathlib
import re
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

ROOT = pathlib.Path(__file__).resolve().parent.parent
FILES = ROOT / "docs" / "files"
OUT = ROOT / "docs" / "data" / "tft-sheets.json"

WORKBOOKS = [
    ("16 Traits", "tft-set-16-traits.xlsx"),
    ("Damage model", "tft-damage-calculator.xlsx"),
]

# Order of the theme palette as a cell's theme="n" indexes it. Excel swaps
# the dark and light pairs relative to how the theme file lists them.
THEME_ORDER = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3",
               "accent4", "accent5", "accent6", "hlink", "folHlink"]


def theme_palette(path):
    """The workbook's twelve theme colours, in cell-index order."""
    with zipfile.ZipFile(path) as z:
        names = [n for n in z.namelist() if n.endswith("theme1.xml")]
        if not names:
            return []
        xml = z.read(names[0]).decode("utf8", "ignore")
    block = re.search(r"<a:clrScheme.*?</a:clrScheme>", xml, re.S)
    if not block:
        return []
    found = dict(
        (name, srgb or sys_last)
        for name, srgb, sys_last in re.findall(
            r'<a:(\w+)>\s*<a:(?:srgbClr val="([0-9A-Fa-f]{6})"'
            r'|sysClr[^>]*lastClr="([0-9A-Fa-f]{6})")',
            block.group(0),
        )
    )
    return [found.get(k) for k in THEME_ORDER]


def apply_tint(hex6, tint):
    """Excel lightens or darkens a theme colour by shifting HLS luminance."""
    if not tint:
        return hex6
    r, g, b = (int(hex6[i:i + 2], 16) / 255 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l * (1 + tint) if tint < 0 else l * (1 - tint) + tint
    r, g, b = colorsys.hls_to_rgb(h, min(1.0, max(0.0, l)), s)
    return "".join(f"{round(v * 255):02X}" for v in (r, g, b))


def argb(colour, palette):
    if colour is None:
        return None
    rgb = getattr(colour, "rgb", None)
    if isinstance(rgb, str) and len(rgb) == 8 and rgb[:2] != "00":
        return "#" + rgb[2:]
    # A theme colour carries an index into the palette plus a tint, and no
    # RGB of its own. Without resolving it the cell renders unpainted.
    theme = getattr(colour, "theme", None)
    if isinstance(theme, int) and 0 <= theme < len(palette) and palette[theme]:
        tint = getattr(colour, "tint", 0) or 0
        return "#" + apply_tint(palette[theme], tint if isinstance(tint, float) else 0)
    return None


def style_of(cell, palette):
    s = {}
    if cell.fill is not None and cell.fill.patternType:
        bg = argb(cell.fill.fgColor, palette)
        if bg and bg.lower() != "#ffffff":
            s["bg"] = bg
    if cell.font is not None:
        fg = argb(cell.font.color, palette)
        if fg and fg.lower() != "#000000":
            s["fg"] = fg
        if cell.font.bold:
            s["b"] = 1
    align = getattr(cell.alignment, "horizontal", None)
    if align in ("center", "right", "left"):
        s["a"] = align
    fmt = cell.number_format or ""
    if "%" in fmt:
        s["p"] = 1                      # render as a percentage
    return s


def dropdown_options(wb, formula):
    """Resolve a list validation into the options it offers.

    The option ranges are themselves formulas pointing at other sheets, so
    they are read from the cached-value copy of the workbook: the names the
    dropdown shows, not the references behind them.
    """
    f = (formula or "").strip()
    if not f:
        return None
    if f.startswith('"') and f.endswith('"'):
        return [x.strip() for x in f[1:-1].split(",") if x.strip()]
    ref = f.lstrip("=").replace("$", "")
    if "!" not in ref:
        return None
    name, rng = ref.split("!", 1)
    name = name.strip("'")
    if name not in wb.sheetnames:
        return None
    src = wb[name]
    c1, r1, c2, r2 = range_boundaries(rng)
    out = []
    for r in range(r1, (r2 or r1) + 1):
        for c in range(c1, (c2 or c1) + 1):
            v = src.cell(row=r, column=c).value
            if v not in (None, ""):
                out.append(str(v))
    return out or None


def validations(wb, ws, values):
    """Map each cell carrying a list validation to its option set.

    The dropdowns are how the workbook is meant to be driven: pick a
    champion, pick items, read the damage. Without them the sheet is a
    picture of a calculator rather than the calculator.
    """
    lists, seen, cells = [], {}, {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        options = dropdown_options(values, dv.formula1)
        if not options:
            continue
        key = " ".join(options)
        if key not in seen:
            seen[key] = len(lists)
            lists.append(options)
        for rng in str(dv.sqref).split():
            c1, r1, c2, r2 = range_boundaries(rng)
            for r in range(r1, (r2 or r1) + 1):
                for c in range(c1, (c2 or c1) + 1):
                    cells[f"{r},{c}"] = seen[key]
    return lists, cells


def conditional_formats(ws, palette):
    """The rules that colour cells from their own values.

    The combat simulation block is entirely conditional: the two HP columns
    run a colour scale from red to green and turn solid red once a fighter
    is dead. Read as static fills the block is blank, which is most of the
    colour the sheet has.
    """
    out = []
    fmt = ws.conditional_formatting
    if not fmt:
        return out
    for group in fmt:
        spans = []
        for ref in str(group.sqref).split():
            c1, r1, c2, r2 = range_boundaries(ref)
            spans.append([r1, c1, r2 or r1, c2 or c1])
        for rule in group.rules:
            base = {"p": rule.priority or 99, "r": spans}
            if rule.type == "cellIs" and rule.dxf is not None:
                fill = rule.dxf.fill
                bg = None
                if fill is not None:
                    bg = argb(fill.bgColor, palette) or argb(fill.fgColor, palette)
                if not bg:
                    continue
                out.append(dict(base, t="cellIs", op=rule.operator,
                                f=[str(x) for x in (rule.formula or [])], bg=bg))
            elif rule.type == "colorScale" and rule.colorScale is not None:
                cs = rule.colorScale
                colours = [argb(c, palette) for c in cs.color]
                if any(c is None for c in colours):
                    continue
                out.append(dict(base, t="colorScale", colors=colours,
                                cfvo=[[v.type, str(v.val) if v.val is not None else None]
                                      for v in cs.cfvo]))
    return out


def sheet_model(ws, styles, index, palette, values):
    merges = []
    covered = set()
    for rng in ws.merged_cells.ranges:
        merges.append([rng.min_row, rng.min_col,
                       rng.max_row - rng.min_row + 1,
                       rng.max_col - rng.min_col + 1])
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))

    cells = {}
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            st = style_of(cell, palette)
            has_value = cell.value not in (None, "")
            if not has_value and not st:
                continue
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
            rec = {}
            v = cell.value
            # Array formulas arrive as an object rather than a string.
            text = getattr(v, "text", None)
            if isinstance(text, str):
                v = text
            if isinstance(v, str) and v.startswith("="):
                rec["f"] = v[1:]
            elif has_value:
                rec["v"] = v if isinstance(v, (int, float, str, bool)) else str(v)
            if st:
                key = json.dumps(st, sort_keys=True)
                if key not in index:
                    index[key] = len(styles)
                    styles.append(st)
                rec["s"] = index[key]
            if rec:
                cells[f"{cell.row},{cell.column}"] = rec

    if not max_row:
        return None

    cols = []
    for c in range(1, max_col + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        w = getattr(dim, "width", None) if dim else None
        cols.append(round((w or 9) * 7.2 + 6))

    lists, dv = validations(values, ws, values)

    return {
        "name": ws.title,
        "rows": max_row,
        "cols": max_col,
        "widths": cols,
        "merges": merges,
        "covered": [f"{r},{c}" for r, c in sorted(covered)],
        "cells": cells,
        "lists": lists,
        "dv": dv,
        "cf": conditional_formats(ws, palette),
    }


def main():
    styles, index = [], {}
    books = []
    for label, fname in WORKBOOKS:
        wb = openpyxl.load_workbook(FILES / fname, data_only=False)
        palette = theme_palette(FILES / fname)
        values = openpyxl.load_workbook(FILES / fname, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            m = sheet_model(ws, styles, index, palette, values)
            if m:
                sheets.append(m)
                formulas = sum(1 for c in m["cells"].values() if "f" in c)
                print(f"  {label} / {ws.title}: {len(m['cells'])} cells, "
                      f"{formulas} formulas, {len(m['dv'])} dropdowns, "
                      f"{len(m['cf'])} conditional rules")
        books.append({"label": label, "file": fname, "sheets": sheets})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"books": books, "styles": styles},
                              ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {OUT} ({OUT.stat().st_size / 1024:.0f} KB, {len(styles)} styles)")


if __name__ == "__main__":
    main()
